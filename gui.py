# === Monitoreo de Tráfico - Waze (GUI) ===
# Cambios en esta versión:
# 1) CRUDOS COMO HTML: la ventana "Crudos" ahora imprime un snippet HTML
#    con <span style="color:..."> para resaltar nombre, tiempo, velocidad y distancia.
# 2) ORDEN EN CLASIFICACIÓN: los logs de clasificación se imprimen por captura
#    en orden: primero INUSUALES (alfabético), luego WATCHLIST/USUALES (alfabético),
#    y finalmente DESCONOCIDOS (alfabético). No se usa jam-level ni estados.
# 3) CONTADOR: se elimina el renglón del log de texto en la ventana principal
#    (se quita el label "📝 Log: ...").
# 4) INTERFAZ: sin cambios de flujo. "Captura instantánea" mantiene lock anti-colisión.
# 5) Compatibilidad: no se asume ningún nombre fijo de funciones en analyzer; se resuelven dinámicamente.

import os, sys, re, time, threading, tempfile, json, logging, logging.handlers
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext
from datetime import datetime, timedelta
from pathlib import Path

# ====== LOG GUI ======
def _ruta_log_usuario(filename: str = "captura_waze_gui.log") -> Path:
    candidatos = []
    if os.name == "nt":
        la = os.environ.get("LOCALAPPDATA")
        if la:
            candidatos.append(Path(la) / "CapturaWaze")
    candidatos.append(Path.home() / ".CapturaWaze")
    candidatos.append(Path(tempfile.gettempdir()) / "CapturaWaze")
    for d in candidatos:
        try:
            d.mkdir(parents=True, exist_ok=True)
            return d / filename
        except Exception:
            pass
    return Path(tempfile.gettempdir()) / filename

LOG_PATH = _ruta_log_usuario("captura_waze_gui.log")
logger = logging.getLogger("captura_waze.gui")
logger.setLevel(logging.INFO)
_detector = None  # singleton del detector

try:
    handler = logging.handlers.RotatingFileHandler(
        LOG_PATH, maxBytes=1_500_000, backupCount=3, encoding="utf-8"
    )
except Exception:
    handler = logging.StreamHandler(stream=sys.stdout)
handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
if not logger.handlers:
    logger.addHandler(handler)
logger.info("GUI iniciada")

# ====== Debug temprano (diagnóstico de arranque) ======
def _early_boot_debug():
    """Diagnóstico de arranque: imprime a consola y al log qué se importó y qué funciones ve la GUI."""
    try:
        print("[GUI] Booting GUI...", flush=True)
    except Exception:
        pass
    try:
        import analyzer as _an_dbg
        print(f"[GUI] analyzer loaded from: {_an_dbg.__file__}", flush=True)
        for n in ("detect_all_segments", "capture_and_save"):
            f = getattr(_an_dbg, n, None)
            print(f"[GUI] analyzer.{n} -> {f}", flush=True)
        globals()["_an_dbg"] = _an_dbg
    except Exception as e:
        print(f"[GUI] analyzer import FAILED: {e}", flush=True)
        try: logger.exception("Analyzer import failed", exc_info=e)
        except Exception: pass

    try:
        import config as _cfg_dbg
        print(f"[GUI] config loaded from: {_cfg_dbg.__file__}", flush=True)
        print(f"[GUI] intervalo_captura (s): {getattr(_cfg_dbg,'intervalo_captura', None)}", flush=True)
        print(f"[GUI] helpers present: "
              f"set_runtime_period_minutes={hasattr(_cfg_dbg,'set_runtime_period_minutes')} | "
              f"get_runtime_period_seconds={hasattr(_cfg_dbg,'get_runtime_period_seconds')} | "
              f"intervalo_captura_sugerido={hasattr(_cfg_dbg,'intervalo_captura_sugerido')}",
              flush=True)
    except Exception as e:
        print(f"[GUI] config import FAILED: {e}", flush=True)

def mostrar_config_inicial(root):
    """
    Dialogo modal de configuración al arrancar:
    - Checkboxes exclusivos: Archivo nuevo (default) / Archivo existente
    - Botón 'Elegir...' abre save/open según selección
    - Modo headless (segundo plano)
    - Modo 24/7 o por horario (HH:MM)
    Persiste en config.json y actualiza las globals de config.
    """
    import os
    import tkinter as tk
    from tkinter import filedialog, ttk, messagebox
    from config import load_cfg, save_cfg, apply_ui_result

    cfg = load_cfg()

    win = tk.Toplevel(root)
    win.title("Configuración inicial")
    win.transient(root)
    win.grab_set()
    win.geometry("560x360")

    # ---------- Archivo ----------
    frm1 = tk.LabelFrame(win, text="Archivo de datos (Excel)", padx=10, pady=8)
    frm1.pack(fill="x", padx=10, pady=6)

    ruta_var = tk.StringVar(value=str(cfg.get("excel_path", "")))

    # Checkboxes exclusivos (mutua exclusión manual)
    usar_nuevo_var = tk.BooleanVar(value=True)   # default: NUEVO
    usar_exist_var = tk.BooleanVar(value=False)

    def _toggle_nuevo():
        # Si se activa 'nuevo', se desactiva 'existente'
        if usar_nuevo_var.get():
            usar_exist_var.set(False)
        else:
            # Evitar quedar con ambos en False: mantener uno activo siempre
            usar_nuevo_var.set(True)

    def _toggle_existente():
        # Si se activa 'existente', se desactiva 'nuevo'
        if usar_exist_var.get():
            usar_nuevo_var.set(False)
        else:
            usar_exist_var.set(True)

    chk_nuevo = tk.Checkbutton(frm1, text="Archivo nuevo", variable=usar_nuevo_var, command=_toggle_nuevo)
    chk_exist = tk.Checkbutton(frm1, text="Archivo existente", variable=usar_exist_var, command=_toggle_existente)
    chk_nuevo.grid(row=0, column=0, sticky="w")
    chk_exist.grid(row=0, column=1, sticky="w", padx=(12,0))

    def _elegir():
        # Botón Elegir... actúa según el checkbox activo
        if usar_exist_var.get():
            p = filedialog.askopenfilename(
                parent=win,
                title="Seleccionar archivo Excel existente",
                filetypes=[("Excel (*.xlsx)", "*.xlsx")]
            )
        else:
            p = filedialog.asksaveasfilename(
                parent=win,
                title="Seleccionar ruta y nombre del archivo nuevo",
                defaultextension=".xlsx",
                filetypes=[("Excel (*.xlsx)", "*.xlsx")],
                initialfile="Captura_Waze.xlsx"
            )
        if p:
            ruta_var.set(p)

    tk.Button(frm1, text="Elegir…", command=_elegir).grid(row=1, column=0, pady=6, sticky="w")
    tk.Label(frm1, textvariable=ruta_var, anchor="w").grid(row=1, column=1, sticky="w", padx=8)

    # ---------- Modo de ejecución ----------
    frm2 = tk.LabelFrame(win, text="Ejecución", padx=10, pady=8)
    frm2.pack(fill="x", padx=10, pady=6)
    headless_var = tk.BooleanVar(value=bool(cfg.get("headless", False)))
    tk.Checkbutton(frm2, text="Segundo plano (Chrome headless)", variable=headless_var).pack(anchor="w")

    # ---------- Modo de operación ----------
    frm3 = tk.LabelFrame(win, text="Horario", padx=10, pady=8)
    frm3.pack(fill="x", padx=10, pady=6)

    modo_var = tk.IntVar(value=1 if bool(cfg.get("modo_247", True)) else 2)
    tk.Radiobutton(frm3, text="24/7", variable=modo_var, value=1).grid(row=0, column=0, sticky="w", pady=(0,4))
    tk.Radiobutton(frm3, text="Horario definido", variable=modo_var, value=2).grid(row=1, column=0, sticky="w")

    def _spin(parent, from_, to_, val):
        sb = tk.Spinbox(parent, from_=from_, to=to_, width=3)
        sb.delete(0, "end"); sb.insert(0, str(val))
        return sb

    row2 = tk.Frame(frm3); row2.grid(row=2, column=0, columnspan=2, pady=6, sticky="w")
    tk.Label(row2, text="Inicio").grid(row=0, column=0, padx=(0,6))
    ini_h = _spin(row2, 0, 23, int(cfg.get("hora_ini_h", 6))); ini_h.grid(row=0, column=1)
    ini_m = _spin(row2, 0, 59, int(cfg.get("hora_ini_m", 0))); ini_m.grid(row=0, column=2, padx=(4,12))

    tk.Label(row2, text="Fin").grid(row=0, column=3, padx=(12,6))
    fin_h = _spin(row2, 0, 23, int(cfg.get("hora_fin_h", 22))); fin_h.grid(row=0, column=4)
    fin_m = _spin(row2, 0, 59, int(cfg.get("hora_fin_m", 0))); fin_m.grid(row=0, column=5, padx=(4,0))

    # ---------- Aceptar ----------
    def aceptar():
        ruta = ruta_var.get().strip()
        if not ruta:
            messagebox.showerror("Falta archivo", "Elige una ruta/archivo Excel.", parent=win)
            return
        # No hacemos nada con el archivo aquí (no crear/abrir). Solo guardamos preferencias.
        nuevo = dict(cfg)
        nuevo["excel_path"] = ruta
        nuevo["headless"] = bool(headless_var.get())
        nuevo["modo_247"] = (modo_var.get() == 1)
        nuevo["hora_ini_h"] = int(ini_h.get())
        nuevo["hora_ini_m"] = int(ini_m.get())
        nuevo["hora_fin_h"] = int(fin_h.get())
        nuevo["hora_fin_m"] = int(fin_m.get())

        save_cfg(nuevo)
        apply_ui_result({
            "modo_247": nuevo["modo_247"],
            "hora_ini": (nuevo["hora_ini_h"], nuevo["hora_ini_m"]),
            "hora_fin": (nuevo["hora_fin_h"], nuevo["hora_fin_m"]),
        })

        win.grab_release()
        win.destroy()

    tk.Button(win, text="Aceptar y continuar", command=aceptar).pack(pady=10)

    # centrar sobre root
    try:
        win.update_idletasks()
        x = root.winfo_x() + (root.winfo_width()//2 - win.winfo_width()//2)
        y = root.winfo_y() + (root.winfo_height()//2 - win.winfo_height()//2)
        win.geometry(f"+{max(0,x)}+{max(0,y)}")
    except Exception:
        pass

    root.wait_window(win)

def create_root():
    """Crea la ventana raíz de Tk y deja trazas claras en consola si algo sale mal."""
    try:
        print("[GUI] Creating Tk root...", flush=True)
        r = tk.Tk()
        r.update_idletasks()
        try:
            r.attributes('-topmost', True)
            r.after(500, lambda: r.attributes('-topmost', False))
        except Exception:
            pass
        print("[GUI] Tk root created OK", flush=True)
        return r
    except Exception as e:
        print(f"[GUI] Tk root creation FAILED: {e}", flush=True)
        raise

# Ejecutar diagnóstico temprano
_early_boot_debug()

def _ensure_detector_started():
    """Arranca (una sola vez) el TrafficDetector con la config persistida."""
    global _detector
    if _detector is not None:
        return _detector

    from analyzer import TrafficDetector
    from models import AppConfig
    from config import load_cfg, WAZE_URL

    cfg_json = load_cfg()
    cfg = AppConfig(
        excel_path=cfg_json.get("excel_path", ""),
        headless=bool(cfg_json.get("headless", False)),
        perfil_persistente=bool(cfg_json.get("perfil_persistente", True)),
        log_level=str(cfg_json.get("log_level", "INFO")),
        waze_url=WAZE_URL,
        modo_247=bool(cfg_json.get("modo_247", True)),
        hora_ini_h=int(cfg_json.get("hora_ini_h", 6)),
        hora_ini_m=int(cfg_json.get("hora_ini_m", 0)),
        hora_fin_h=int(cfg_json.get("hora_fin_h", 22)),
        hora_fin_m=int(cfg_json.get("hora_fin_m", 0)),
        periodicidad_min=int(cfg_json.get("periodicidad_min", 10)),
    )

    det = TrafficDetector(cfg, logger)
    det.start()  # ← abre Chrome, navega a Waze y activa Traffic View
    _detector = det
    return det

def _refresh_browser(motivo: str = ""):
    """
    Refresca el navegador (driver.refresh + esperar panel) usando el detector singleton.
    Es usada 1 min antes de cada captura programada y en la captura instantánea.
    """
    try:
        det = _ensure_detector_started()
        log_clasificacion(f"🔄 Recargando página {('(' + motivo + ')') if motivo else ''}…")
        det.refresh()
        log_clasificacion("✅ Página recargada.")
    except Exception as e:
        log_clasificacion(f"⚠️ No se pudo recargar: {e}")

def _bootstrap_startup():
    """
    1) Asegura que exista un Excel válido (crea uno si falta).
    2) Abre Chrome/Waze levantando el detector.
    """
    # 1) Excel
    from config import load_cfg, save_cfg
    from storage import set_workbook, safe_save_workbook
    import openpyxl, os

    cfg = load_cfg()
    ruta = cfg.get("excel_path", "")

    def _poner_encabezados(ws):
        if ws.max_row == 1 and all(c.value is None for c in ws[1]):
            ws.append(["Fecha","Hora","Tramo","Tiempo (MM:SS)","Tiempo (s)","Velocidad (km/h)","Distancia (km)"])

    if not ruta or not os.path.exists(ruta):
        ruta = filedialog.asksaveasfilename(
            title="Crear archivo Excel para las capturas",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Captura_Waze.xlsx"
        )
        if not ruta:
            messagebox.showwarning("Configuración", "No seleccionaste un Excel. Podrás crear uno más tarde, pero no se guardarán capturas.")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tráfico inusual"
            _poner_encabezados(ws)
            safe_save_workbook(wb, ruta)
            set_workbook(wb, ruta)
            cfg["excel_path"] = ruta
            save_cfg(cfg)
    else:
        try:
            wb = openpyxl.load_workbook(ruta)
        except Exception as e:
            messagebox.showerror("Excel", f"No se pudo abrir el Excel configurado:\n{e}")
        else:
            if "Tráfico inusual" not in wb.sheetnames:
                ws = wb.create_sheet("Tráfico inusual")
                _poner_encabezados(ws)
            safe_save_workbook(wb, ruta)
            set_workbook(wb, ruta)

    # 2) Arrancar Chrome/Waze
    try:
        _ensure_detector_started()
    except Exception as e:
        messagebox.showerror("Navegador", f"No se pudo iniciar el navegador con Selenium:\n{e}")

# ====== IMPORTS del proyecto ======
from storage import ruta_log_txt, safe_save_workbook  # log y guardado seguro
from config import (
    modo_247, hora_ini, hora_fin, intervalo_captura_sugerido,
    set_runtime_period_minutes, get_runtime_period_seconds,
    esta_dentro_horario, proximo_inicio_desde, alinear_a_intervalo
)
try:
    from storage import wb, archivo_excel
except Exception:
    wb = None
    archivo_excel = None

# Resolver funciones del analyzer sin conocer sus nombres exactos
import analyzer as _an

def _resolve_fn(possible_names):
    for name in possible_names:
        fn = getattr(_an, name, None)
        if callable(fn):
            logger.info(f"[GUI] Usando analyzer.{name}()")
            return fn
    raise ImportError(f"No se encontró ninguna de estas funciones en analyzer.py: {possible_names}")

# Detectar/extract crudos (sin guardar)
analyzer_detect_all = _resolve_fn([
    "detect_all_segments", "detect_segments", "extract_segments",
    "capturar_todos_los_tramos", "extraer_tramos", "listar_tramos", "scan_segments"
])

# Capturar+guardar (retorna tuple: guardados, usuales, inusuales)
analyzer_capture_and_save = _resolve_fn([
    "capture_and_save", "capturar_y_guardar", "procesar_y_guardar_tramos",
    "run_capture", "save_capture"
])

# ====== Estado global GUI ======
stop_event = threading.Event()
captura_lock = threading.Lock()
siguiente_captura = None
refrescado_este_ciclo = False

# ====== Ventanas de logs ======
# Clasificación
vista_cls_activa = False
win_cls = None
text_cls = None

# Crudos
vista_raw_activa = False
win_raw = None
text_raw = None

def _append_to_text(text_widget, line: str):
    try:
        text_widget.insert("end", line + "\n")
        text_widget.see("end")
    except Exception:
        pass

# -------- Clasificación (ordenada por captura) --------
def abrir_logs_clasificacion():
    global vista_cls_activa, win_cls, text_cls
    if vista_cls_activa and win_cls:
        try:
            win_cls.lift(); return
        except Exception:
            vista_cls_activa = False
    vista_cls_activa = True
    win_cls = tk.Toplevel(root)
    win_cls.title("Clasificación — Logs de captura (Inusual → Watchlist → Desconocidos)")
    win_cls.geometry("900x520")
    text_cls = scrolledtext.ScrolledText(win_cls, wrap="word", font=("Consolas", 10))
    text_cls.pack(fill="both", expand=True, padx=8, pady=8)
    frm_btn = tk.Frame(win_cls); frm_btn.pack(fill="x", padx=8, pady=(0,8))
    tk.Button(frm_btn, text="Limpiar", command=lambda: text_cls.delete("1.0","end")).pack(side="left")
    def copiar():
        data = text_cls.get("1.0", "end-1c")
        win_cls.clipboard_clear(); win_cls.clipboard_append(data)
    tk.Button(frm_btn, text="Copiar todo", command=copiar).pack(side="left", padx=6)
    def on_close():
        global vista_cls_activa, win_cls, text_cls
        vista_cls_activa = False; text_cls = None
        try: win_cls.destroy()
        except Exception: pass
        win_cls = None
    win_cls.protocol("WM_DELETE_WINDOW", on_close)

def log_clasificacion(mensaje: str):
    logger.info(mensaje)
    if not vista_cls_activa or text_cls is None:
        return
    ts = datetime.now().strftime("%H:%M:%S")
    root.after(0, lambda: _append_to_text(text_cls, f"[{ts}] {mensaje}"))

# -------- Crudos --------
def abrir_logs_crudos():
    global vista_raw_activa, win_raw, text_raw
    if vista_raw_activa and win_raw:
        try:
            win_raw.lift(); return
        except Exception:
            vista_raw_activa = False
    vista_raw_activa = True
    win_raw = tk.Toplevel(root)
    win_raw.title("Crudos — Snippet HTML (pre-clasificación)")
    win_raw.geometry("900x520")
    text_raw = scrolledtext.ScrolledText(win_raw, wrap="word", font=("Consolas", 10))
    text_raw.pack(fill="both", expand=True, padx=8, pady=8)
    frm_btn = tk.Frame(win_raw); frm_btn.pack(fill="x", padx=8, pady=(0,8))
    tk.Button(frm_btn, text="Limpiar", command=lambda: text_raw.delete("1.0","end")).pack(side="left")
    def copiar():
        data = text_raw.get("1.0", "end-1c")
        win_raw.clipboard_clear(); win_raw.clipboard_append(data)
    tk.Button(frm_btn, text="Copiar todo", command=copiar).pack(side="left", padx=6)
    def on_close():
        global vista_raw_activa, win_raw, text_raw
        vista_raw_activa = False; text_raw = None
        try: win_raw.destroy()
        except Exception: pass
        win_raw = None
    win_raw.protocol("WM_DELETE_WINDOW", on_close)

# === generador de snippet HTML de crudo con resaltado por color ===

# --- helpers de coloreo para el Text de "Crudos" ---

def _ensure_text_tags(widget):
    """Define (una sola vez) las etiquetas de color usadas en el Text."""
    try:
        # Si ya existe alguna, asumimos que todas están
        widget.tag_cget("c_blue", "foreground")
        return
    except Exception:
        pass
    widget.tag_config("c_blue",   foreground="#1E40AF")  # nombre
    widget.tag_config("c_orange", foreground="#EA580C")  # distancia
    widget.tag_config("c_green",  foreground="#16A34A")  # tiempo (min)
    widget.tag_config("c_purple", foreground="#7C3AED")  # velocidad (km/h")


def _apply_color(text_widget, line_start_index, pattern, tag_name):
    """
    Colorea todas las coincidencias de 'pattern' (regex) **solo en esa línea**,
    calculando offsets con '+Nc' desde el inicio de la línea.
    """
    import re as _re

    # Texto de esa línea nada más
    line_text = text_widget.get(line_start_index, f"{line_start_index} lineend")

    for m in _re.finditer(pattern, line_text):
        a, b = m.span()
        # Índices seguros usando offsets sobre el linestart
        start_idx = f"{line_start_index}+{a}c"
        end_idx   = f"{line_start_index}+{b}c"
        try:
            text_widget.tag_add(tag_name, start_idx, end_idx)
        except Exception:
            # No dejamos que un fallo en un match interrumpa el resto
            pass


def _insert_crudo_with_colors(item: dict):
    """
    Inserta UNA línea HTML (mock) en la vista 'Crudos' y aplica color a:
      - Nombre (azul)
      - Distancia (naranja)
      - Tiempo (min) actual (verde)
      - Velocidad (km/h) actual (morado)
    Usa índices 'end-1c linestart' + offsets para evitar TclError.
    """
    from datetime import datetime
    import html

    if text_raw is None:
        return

    _ensure_text_tags(text_raw)

    # Campos
    nombre = (item.get("nombre") or item.get("name") or "").strip()
    dist   = (item.get("dist_raw") or item.get("dist") or "").strip()
    curr   = (item.get("current_raw") or item.get("current") or "").strip()
    hist   = (item.get("historic_raw") or item.get("historic") or "").strip()

    # Escapar por seguridad (aunque es mock HTML)
    nombre_e = html.escape(nombre)
    dist_e   = html.escape(dist)
    curr_e   = html.escape(curr)
    hist_e   = html.escape(hist)

    ts = datetime.now().strftime("%H:%M:%S")

    # Línea “HTML” (como la que me mostraste). Pongo spans con estilos,
    # pero el color real lo damos con tags del Text (esto es “decorativo”).
    line = (
        f"[{ts}] "
        f"<app-traffic-view-route><app-traffic-view-sidebar-section>"
        f"<div class=\"route-info-container\">"
        f"<div class=\"basic-info\">"
        f"<wz-subhead4><span style=\"color:#1E40AF\">{nombre_e}</span></wz-subhead4>"
        f"<wz-caption class=\"route-distance\"><span style=\"color:#EA580C\">{dist_e}</span></wz-caption>"
        f"</div>"
        f"<div class=\"traffic-info\"><app-traffic-view-route-stats>"
        f"<wz-caption class=\"current-stat\"><span style=\"color:#16A34A\">{curr_e.split('|')[0].strip()}</span>"
        f"{(' | <span style=\"color:#7C3AED\">' + curr_e.split('|')[1].strip() + '</span>') if '|' in curr_e else ''}"
        f"</wz-caption>"
        f"<wz-caption class=\"historic-stat\">{hist_e}</wz-caption>"
        f"</app-traffic-view-route-stats></div>"
        f"<div class=\"route-menu\"><wz-menu><wz-menu-item> Edit </wz-menu-item>"
        f"<wz-menu-item> Delete route </wz-menu-item></wz-menu></div>"
        f"</div></app-traffic-view-sidebar-section></app-traffic-view-route>\n"
    )

    # Insertar al final
    text_raw.insert("end", line)

    # Inicio de la línea recién insertada (índice robusto)
    line_start = text_raw.index("end-1c linestart")

    # Patrones a resaltar (solo en ESTA línea):
    # 1) Nombre dentro del subhead (azul)
    if nombre:
        _apply_color(
            text_raw,
            line_start,
            rf"<wz-subhead4>\s*{re.escape(nombre_e)}\s*|<wz-subhead4><span[^>]*>{re.escape(nombre_e)}</span>\s*</wz-subhead4>",
            "c_blue",
        )

    # 2) Distancia (naranja): número + ' km'
    #    Tomamos la cifra dentro de route-distance, o cualquier 'X.xx km' de la línea
    _apply_color(
        text_raw,
        line_start,
        r"\b\d+(?:[.,]\d+)?\s*km\b",
        "c_orange",
    )

    # 3) Tiempo actual (verde): "<n> min"
    _apply_color(
        text_raw,
        line_start,
        r"\b\d+\s*min(?:utos)?\b",
        "c_green",
    )

    # 4) Velocidad actual (morado): "##.## km/h"
    _apply_color(
        text_raw,
        line_start,
        r"\b\d+(?:[.,]\d+)?\s*km/?h\b",
        "c_purple",
    )

    # Desplazar el scroll al final
    try:
        text_raw.see("end")
    except Exception:
        pass

def log_crudo(item: dict):
    if not vista_raw_activa or text_raw is None:
        return
    root.after(0, lambda: _insert_crudo_with_colors(item))


# ====== UI helpers ======
def _fmt_restante(segundos: int) -> str:
    if segundos < 0: segundos = 0
    h, rem = divmod(segundos, 3600); m, s = divmod(rem, 60)
    return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"

def _post_captura_ui(guardados: int, usuales: int, inusuales: int, desconocidos: int):
    try:
        lbl_resultado_normal.config(text=f"📘 Guardados: {guardados}")
        # Se elimina el label del log de texto (requisito #4).
        lbl_usuales_inusuales.config(
            text=f"✅ Usuales: {usuales}   ⚠️ Inusuales: {inusuales}   ❓ Desconocidos: {desconocidos}"
        )
    except Exception:
        pass

# ====== Clasificación desde crudos (sin estados/jam-level) ======
def _clasif_from_raw(item: dict):
    """
    True -> USUAL (Watchlist)
    False -> INUSUAL (sección Unusual traffic)
    None -> Desconocido (sin flag)
    """
    flag = (item.get("section_flag") or item.get("section") or "").strip().lower()
    if flag == "unusual":
        return False
    if flag == "watch":
        return True
    if "es_usual" in item:
        v = item.get("es_usual")
        if v is True: return True
        if v is False: return False
    return None

def _emitir_crudos_y_clasificacion(tramos):
    """
    Muestra crudos (HTML) y luego imprime una CLASIFICACIÓN ORDENADA por captura:
      1) INUSUALES (alfabético por nombre)
      2) WATCHLIST/USUALES (alfabético por nombre)
      3) DESCONOCIDOS (alfabético por nombre)
    """
    if not tramos:
        return

    # Primero, enviar CRUDOS (uno por uno, en el orden recibido)
    for t in tramos:
        log_crudo(t)

    # Luego, clasificación ordenada
    inus = []
    usua = []
    desc = []
    for t in tramos:
        es = _clasif_from_raw(t)
        nombre = (t.get("nombre") or t.get("name") or "").strip()
        dist = (t.get("dist_raw") or t.get("dist") or "").strip()
        curr = (t.get("current_raw") or t.get("current") or "").strip()
        hist = (t.get("historic_raw") or t.get("historic") or "").strip()
        visible = curr or hist
        linea_base = f"{nombre} | {visible} | {dist}"
        if es is False:
            inus.append(linea_base)
        elif es is True:
            usua.append(linea_base)
        else:
            desc.append(linea_base)

    def _ordenado(lista):
        return sorted(lista, key=lambda s: s.split(" | ", 1)[0].lower())

    inus_sorted = _ordenado(inus)
    usua_sorted = _ordenado(usua)
    desc_sorted = _ordenado(desc)

    if inus_sorted:
        log_clasificacion("── Inusuales (alfabético) ──")
        for x in inus_sorted:
            log_clasificacion("[INUSUAL] " + x)
    if usua_sorted:
        log_clasificacion("── Watchlist (alfabético) ──")
        for x in usua_sorted:
            log_clasificacion("[USUAL] " + x)
    if desc_sorted:
        log_clasificacion("── Desconocidos (alfabético) ──")
        for x in desc_sorted:
            log_clasificacion("[?] " + x)

# ====== Captura principal ======
def _captura_ejecucion():
    # Asegura que el detector (y Chrome) están levantados
    det = _ensure_detector_started()

    # 1) EXTRAER crudos (sin guardar) → imprimir crudos y clasificación ordenada
    tramos_crudos = det.detect_all()
    _emitir_crudos_y_clasificacion(
        [t.__dict__ if hasattr(t, "__dict__") else t for t in tramos_crudos]
    )

    # Conteo de desconocidos
    try:
        desconocidos = sum(1 for t in tramos_crudos if getattr(t, "es_usual", None) is None)
    except Exception:
        desconocidos = 0

    # 2) Guardar en Excel y devolver contadores
    from storage import guardar_tramos
    guardados, cont_usuales, cont_inusuales = guardar_tramos(tramos_crudos)

    return guardados, cont_usuales, cont_inusuales, desconocidos

def captura_instantanea():
    if captura_lock.locked():
        messagebox.showinfo("Captura en curso", "Ya hay una captura ejecutándose.")
        return
    def _run():
        with captura_lock:
            try:
                _refresh_browser("instantánea")
                log_clasificacion("▶️ Captura instantánea solicitada")
                g, u, i, d = _captura_ejecucion()
                _post_captura_ui(g, u, i, d)
                log_clasificacion(f"✔️ Instantánea lista. Guardados={g} (Usuales={u}, Inusuales={i}, Desconocidos={d})")
            except Exception as e:
                log_clasificacion(f"❌ Error en captura instantánea: {e}")
    threading.Thread(target=_run, daemon=True).start()

# ====== Hilo del ciclo ======
def ciclo():
    global siguiente_captura, refrescado_este_ciclo
    while not stop_event.is_set():
        try:
            ahora = datetime.now()
            if not esta_dentro_horario(ahora):
                siguiente_captura = alinear_a_intervalo(
                    proximo_inicio_desde(ahora), get_runtime_period_seconds()
                )
                refrescado_este_ciclo = False
                time.sleep(1)
                continue

            if siguiente_captura is None:
                siguiente_captura = alinear_a_intervalo(ahora, get_runtime_period_seconds())

            restante = (siguiente_captura - ahora).total_seconds()

            # 🔄 Recarga 1 minuto antes de la captura programada (solo una vez por ciclo)
            if 0 < restante <= 60 and not refrescado_este_ciclo:
                refrescado_este_ciclo = True
                _refresh_browser("ciclo programado (t-60s)")

            if ahora >= siguiente_captura:
                if not captura_lock.locked():
                    with captura_lock:
                        try:
                            log_clasificacion("⏱️ Ejecutando captura programada…")
                            g, u, i, d = _captura_ejecucion()
                            _post_captura_ui(g, u, i, d)
                        except Exception as e:
                            log_clasificacion(f"⚠️ Error en ciclo: {e}")
                siguiente_captura = ahora + timedelta(seconds=get_runtime_period_seconds())
                refrescado_este_ciclo = False

            time.sleep(1)
        except Exception as e:
            log_clasificacion(f"Loop warning: {e}")
            time.sleep(2)


# ====== Ventana principal ======
root = create_root()
root.title("Monitoreo de Tráfico - Waze")
mostrar_config_inicial(root)

lbl_contador = tk.Label(root, text="⏳ Próxima captura en: --", font=("Arial", 14))
lbl_contador.pack(pady=10)

lbl_resultado_normal = tk.Label(root, text="📘 Guardados: --", font=("Arial", 12))
lbl_resultado_normal.pack()

# (Requisito #4) Se elimina el label del log de texto.
lbl_usuales_inusuales = tk.Label(root, text="✅ Usuales: --   ⚠️ Inusuales: --   ❓ Desconocidos: --", font=("Arial", 12))
lbl_usuales_inusuales.pack(pady=4)

# Botones
frm_btns = tk.Frame(root); frm_btns.pack(pady=(6,10))
btn_shot = tk.Button(frm_btns, text="Captura instantánea", command=captura_instantanea,
                     bg="#2463EB", fg="white", activebackground="#1E4ECC", activeforeground="white")
btn_shot.pack(side="left", padx=(0,10))

def detener():
    stop_event.set()
    try:
        if wb and archivo_excel:
            try:
                safe_save_workbook(wb, archivo_excel)
            except Exception:
                pass
    except Exception:
        pass

    try:
        from analyzer import shutdown_detectors
        shutdown_detectors()
    except Exception:
        pass

    try:
        root.destroy()
    except Exception:
        pass

btn_detener = tk.Button(frm_btns, text="Detener Captura", command=detener, bg="red", fg="white")
btn_detener.pack(side="left")

# Logs checkboxes
frm_logs = tk.Frame(root); frm_logs.pack(pady=(0, 6))
var_cls = tk.BooleanVar(value=True)
var_raw = tk.BooleanVar(value=False)

def _toggle_cls():
    if var_cls.get(): abrir_logs_clasificacion()
    else:
        global vista_cls_activa, win_cls, text_cls
        if vista_cls_activa and win_cls:
            try: win_cls.destroy()
            except Exception: pass
        vista_cls_activa, win_cls, text_cls = False, None, None

def _toggle_raw():
    if var_raw.get(): abrir_logs_crudos()
    else:
        global vista_raw_activa, win_raw, text_raw
        if vista_raw_activa and win_raw:
            try: win_raw.destroy()
            except Exception: pass
        vista_raw_activa, win_raw, text_raw = False, None, None

tk.Checkbutton(frm_logs, text="Mostrar clasificación (logs)", variable=var_cls, command=_toggle_cls).pack(side="left")
tk.Checkbutton(frm_logs, text="Mostrar crudos (pre-clasificación)", variable=var_raw, command=_toggle_raw).pack(side="left", padx=12)

# Periodicidad
frm_per_live = tk.LabelFrame(root, text="Periodicidad", padx=10, pady=6)
frm_per_live.pack(pady=(0,10))
from config import intervalo_captura_sugerido, set_runtime_period_minutes, get_runtime_period_seconds, esta_dentro_horario, proximo_inicio_desde, alinear_a_intervalo
var_per_live = tk.IntVar(value=intervalo_captura_sugerido // 60 if intervalo_captura_sugerido else 10)
def _aplicar_per_live():
    mins = int(var_per_live.get())
    set_runtime_period_minutes(max(10, mins))
    ahora = datetime.now()
    global siguiente_captura
    if esta_dentro_horario(ahora):
        siguiente_captura = alinear_a_intervalo(ahora, get_runtime_period_seconds())
    messagebox.showinfo("Periodicidad actualizada", f"Captura cada {mins} minutos.")
for mins in (10, 15, 30, 60):
    tk.Radiobutton(frm_per_live, text=f"{mins} min", variable=var_per_live, value=mins,
                   command=_aplicar_per_live).pack(side="left", padx=6)

# Abre por defecto la ventana de Clasificación
_toggle_cls()

_bootstrap_startup()



# Primera captura programada
siguiente_captura = alinear_a_intervalo(
    proximo_inicio_desde(datetime.now()) if not esta_dentro_horario(datetime.now()) else datetime.now(),
    get_runtime_period_seconds()
)

# Actualizador del contador
def actualizar_contador():
    if stop_event.is_set():
        return
    ahora = datetime.now()
    if not esta_dentro_horario(ahora):
        proximo = alinear_a_intervalo(proximo_inicio_desde(ahora), get_runtime_period_seconds())
        restante = int((proximo - ahora).total_seconds())
        lbl_contador.config(
            text=f"⏸️ Pausa por horario — próxima ventana: {proximo.strftime('%H:%M')} (en {restante//60:02d}:{restante%60:02d})"
        )
    else:
        if siguiente_captura is None:
            lbl_contador.config(text="⏳ Próxima captura en: --")
        else:
            restante = int((siguiente_captura - ahora).total_seconds())
            lbl_contador.config(text=f"⏳ Próxima captura en: {_fmt_restante(restante)}")
    root.after(1000, actualizar_contador)
actualizar_contador()

# Hilo del ciclo
threading.Thread(target=ciclo, daemon=True).start()

# Cierre limpio
def _on_close():
    detener()
root.protocol("WM_DELETE_WINDOW", _on_close)
root.mainloop()
print("[GUI] mainloop ended", flush=True)
