# storage.py
# Cambios en esta versión:
# 1) GUARDADO por categorías:
#    - Hoja 1: "Tráfico inusual" (siempre al inicio).
#    - Hojas siguientes: UNA por cada tramo de Watchlist (USUAL), ordenadas alfabéticamente.
#    - "Desconocidos" (si aparece) se ubica al FINAL, fuera del bloque alfabético de Watchlist.
# 2) ORDENACIÓN de hojas tras guardar: se reordena el índice de hojas para mantener
#    el criterio anterior (inusual primero, luego watchlist alfabético, luego desconocidos).
# 3) Creación de hoja Watchlist si no existe, insertándola en la posición correcta
#    para conservar el orden alfabético.

from __future__ import annotations

import os
import time
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional, Iterable
import json

def _wal_paths_for(excel_path: str) -> tuple[Path, Path]:
    """
    Calcula rutas de WAL y del marcador de commit para un Excel dado.
    Usamos: <archivo>.xlsx.wal.jsonl  y  <archivo>.xlsx.wal.jsonl.commit
    """
    p = Path(excel_path)
    wal = Path(str(p) + ".wal.jsonl")
    mark = Path(str(wal) + ".commit")
    try:
        wal.parent.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass
    return wal, mark

def _wal_write_transaction(excel_path: str, tramos_norm: list[dict]) -> None:
    """Escribe la transacción como NDJSON (1 tramo por línea) y fuerza a disco."""
    wal, _ = _wal_paths_for(excel_path)
    with open(wal, "w", encoding="utf-8") as f:
        for t in tramos_norm:
            json.dump(t, f, ensure_ascii=False)
            f.write("\n")
        f.flush()
        try:
            os.fsync(f.fileno())
        except Exception:
            pass

def _wal_mark_committed(excel_path: str) -> None:
    """Crea un marcador de commit; si existe, lo actualiza (touch)."""
    _, mark = _wal_paths_for(excel_path)
    try:
        with open(mark, "w", encoding="utf-8") as f:
            f.write("ok")
        try:
            os.fsync(f.fileno())
        except Exception:
            pass
    except Exception:
        pass

def _wal_clear(excel_path: str) -> None:
    """Elimina WAL y el marcador de commit."""
    wal, mark = _wal_paths_for(excel_path)
    for p in (wal, mark):
        try:
            if p.exists():
                p.unlink()
        except Exception:
            pass

def recover_from_wal() -> None:
    """
    Si existe WAL sin commit → re-aplica los tramos pendientes al Excel.
    Si existe WAL con commit → limpia ambos (fue guardado previamente).
    """
    global wb, archivo_excel
    if not archivo_excel:
        return
    wal, mark = _wal_paths_for(archivo_excel)
    if not wal.exists():
        return

    # Si hay commit, significa que ya se alcanzó a guardar y faltó limpiar
    if mark.exists():
        _wal_clear(archivo_excel)
        return

    # Reaplicar
    tramos = []
    try:
        with open(wal, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    tramos.append(json.loads(line))
                except Exception:
                    continue
    except Exception:
        return

    if not tramos:
        _wal_clear(archivo_excel)
        return

    # Guardar sin re-escribir WAL
    try:
        guardar_tramos(tramos, _skip_wal=True)
    finally:
        _wal_clear(archivo_excel)

# (Opcional) openpyxl solo para tipado; no se importa nada pesado aquí
try:
    import openpyxl  # noqa: F401
except Exception:
    pass

# === Estado global mínimo (otros módulos pueden establecerlos) ===
wb = None  # type: ignore  # openpyxl.Workbook
archivo_excel: Optional[str] = None

def set_workbook(workbook, excel_path: str) -> None:
    """Permite a otros módulos registrar el workbook y su ruta."""
    global wb, archivo_excel
    wb = workbook
    archivo_excel = excel_path

# === Helpers de rutas para logs/archivos ===
def _dir_escribible(d: Path) -> bool:
    try:
        d.mkdir(parents=True, exist_ok=True)
        p = d / (".writetest_" + str(os.getpid()))
        with open(p, "w", encoding="utf-8") as f:
            f.write("ok")
        try:
            p.unlink()
        except Exception:
            pass
        return True
    except Exception:
        return False

def _ruta_log_usuario(filename: str = "captura_waze.log") -> Path:
    """Devuelve una ruta válida y escribible para archivos de log."""
    candidatos = []
    if os.name == "nt":
        la = os.environ.get("LOCALAPPDATA")
        if la:
            candidatos.append(Path(la) / "CapturaWaze")
    candidatos.append(Path.home() / ".CapturaWaze")
    candidatos.append(Path(tempfile.gettempdir()) / "CapturaWaze")
    for d in candidatos:
        if _dir_escribible(d):
            return d / filename
    return Path(tempfile.gettempdir()) / filename

def ruta_log_txt() -> str:
    """
    Ruta del log 'plano' diario. Si hay un Excel asociado, usa su nombre como base.
    Ej.: Captura_Waze__LOG_YYYYMMDD.txt
    """
    try:
        log_dir = _ruta_log_usuario().parent
        log_dir.mkdir(parents=True, exist_ok=True)
    except Exception:
        log_dir = Path(tempfile.gettempdir()) / "CapturaWaze"

    base_name = "Captura_Waze"
    try:
        if archivo_excel:
            base_name = Path(os.path.splitext(archivo_excel)[0]).name
    except Exception:
        pass

    fecha = datetime.now().strftime("%Y%m%d")
    return str((log_dir / f"{base_name}__LOG_{fecha}.txt").resolve())

# === Guardado robusto de Excel ===
def atomic_save_workbook(workbook, ruta: str) -> str:
    """Guardado atómico: escribe a un archivo temporal y reemplaza. Devuelve la ruta final."""
    if not ruta:
        raise ValueError("Se requiere 'ruta' para guardar el workbook.")
    dir_ = os.path.dirname(ruta) or "."
    fd, tmp = tempfile.mkstemp(prefix="~waze_", suffix=".xlsx", dir=dir_)
    os.close(fd)
    try:
        workbook.save(tmp)
        os.replace(tmp, ruta)
        return ruta
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except Exception:
            pass

def safe_save_workbook(workbook, ruta: str, intentos: int = 5, espera_s: float = 1.5) -> str:
    """
    Intenta guardar con reintentos (útil si el archivo está brevemente bloqueado).
    No abre diálogos; si falla definitivamente, relanza la última excepción.
    """
    last_exc = None
    for i in range(intentos):
        try:
            return atomic_save_workbook(workbook, ruta)
        except Exception as e:  # PermissionError, etc.
            last_exc = e
            if i < intentos - 1:
                time.sleep(espera_s)
            else:
                raise e
    if last_exc:
        raise last_exc
    return ruta

# === Utilidades internas ===
def _poner_encabezados(ws):
    if ws.max_row == 1 and all((cell.value is None) for cell in ws[1]):
        ws.append([
            "Fecha", "Hora", "Tramo",
            "Tiempo (MM:SS)", "Tiempo (s)",
            "Velocidad (km/h)", "Distancia (km)"
        ])

def _ordenar_hojas_watchlist():
    """
    Reordena las hojas del workbook:
      [0] "Tráfico inusual"
      [1..N] todas las hojas de watchlist (todas salvo "Tráfico inusual" y "Desconocidos"), orden alfabético
      [última] "Desconocidos" (si existe)
    """
    if wb is None:
        return
    sheets = list(wb.sheetnames)
    inus_name = "Tráfico inusual"
    desc_name = "Desconocidos"

    # Garantizar que la de inusual exista
    if inus_name not in sheets:
        ws = wb.create_sheet(inus_name, 0)
        _poner_encabezados(ws)
        sheets = list(wb.sheetnames)

    # Separar categorías
    watchlist_sheets = [s for s in sheets if s not in (inus_name, desc_name)]
    watchlist_sheets.sort(key=lambda x: x.lower())

    new_order = [inus_name] + watchlist_sheets
    if desc_name in sheets:
        new_order.append(desc_name)

    # Asignar nuevo orden
    wb._sheets = [wb[s] for s in new_order]  # openpyxl internals, aceptado para reordenar
    # Nota: no cambiamos titles; solo reordenamos.

def _insertar_hoja_watchlist_en_posicion(title: str):
    """
    Crea una hoja de watchlist con 'title' si no existe, insertándola en la
    posición correcta para mantener el orden alfabético (después de "Tráfico inusual"
    y antes de "Desconocidos").
    """
    if wb is None:
        return
    inus_name = "Tráfico inusual"
    desc_name = "Desconocidos"

    if title in wb.sheetnames:
        return  # ya existe

    # Construir lista ordenada de destino
    existing = [s for s in wb.sheetnames if s not in (inus_name, desc_name)]
    existing.append(title)
    existing.sort(key=lambda x: x.lower())

    # La posición final debe ser: 0 inusual, luego índice 1.. en el orden de existing
    target_index = 1 + existing.index(title)

    # Crear la hoja al final y luego moverla a la posición deseada
    ws = wb.create_sheet(title)
    _poner_encabezados(ws)

    # Reordenar moviendo el objeto
    current = list(wb._sheets)
    current.remove(ws)
    current.insert(target_index, ws)

    # Mantener "Desconocidos" (si existe) al final
    if desc_name in wb.sheetnames:
        ws_desc = wb[desc_name]
        current = [s for s in current if s is not ws_desc] + [ws_desc]

    # Asegurar que "Tráfico inusual" esté primero
    ws_inus = wb[inus_name]
    current = [s for s in current if s is not ws_inus]
    current.insert(0, ws_inus)

    wb._sheets = current

# === API de guardado ===
def guardar_tramos(tramos) -> tuple[int, int, int]:
    """
    Guarda:
      - INUSUAL   (es_usual is False) -> hoja "Tráfico inusual" [siempre en índice 0]
      - USUAL     (es_usual is True)  -> hoja propia por tramo (orden alfabético global)
      - DESCONOC. (es_usual is None)  -> hoja "Desconocidos" [se ubica al final si existe]
    Devuelve: (total_guardados, total_usuales, total_inusuales)
    """
    if wb is None or not archivo_excel:
        return 0, 0, 0

    from models import nombre_hoja_seguro
    from openpyxl.utils.exceptions import IllegalCharacterError

    # Asegurar hojas base
    hoja_inus = wb["Tráfico inusual"] if "Tráfico inusual" in wb.sheetnames else wb.create_sheet("Tráfico inusual", 0)
    _poner_encabezados(hoja_inus)

    # "Desconocidos" se crea solo si aparece alguno
    hoja_desc = wb["Desconocidos"] if "Desconocidos" in wb.sheetnames else None
    if hoja_desc:
        _poner_encabezados(hoja_desc)

    guardados = u = i = 0
    ahora = datetime.now()

    def _val(obj, k, default=None):
        if isinstance(obj, dict):
            return obj.get(k, default)
        return getattr(obj, k, default)

    for t in (tramos or []):
        nombre = (_val(t, "nombre") or "").strip()
        tiempo_mmss = _val(t, "tiempo_mmss") or ""
        tiempo_seg  = _val(t, "tiempo_seg")
        vel_kmh     = _val(t, "vel_kmh")
        dist_km     = _val(t, "dist_km")
        es_usual    = _val(t, "es_usual")  # True / False / None

        try:
            if es_usual is True:
                # USUAL → hoja por tramo (orden alfabético en pestañas)
                hoja = nombre_hoja_seguro(nombre)
                if hoja not in wb.sheetnames:
                    _insertar_hoja_watchlist_en_posicion(hoja)
                ws = wb[hoja]
                _poner_encabezados(ws)
                ws.append([
                    ahora.strftime("%d/%m/%Y"),
                    ahora.strftime("%H:%M:%S"),
                    nombre,
                    tiempo_mmss or "",
                    int(tiempo_seg) if isinstance(tiempo_seg, (int, float)) else "",
                    vel_kmh if vel_kmh is not None else "",
                    dist_km if dist_km is not None else ""
                ])
                u += 1
                guardados += 1

            elif es_usual is False:
                # INUSUAL → "Tráfico inusual"
                hoja_inus.append([
                    ahora.strftime("%d/%m/%Y"),
                    ahora.strftime("%H:%M:%S"),
                    nombre,
                    tiempo_mmss or "",
                    int(tiempo_seg) if isinstance(tiempo_seg, (int, float)) else "",
                    vel_kmh if vel_kmh is not None else "",
                    dist_km if dist_km is not None else ""
                ])
                i += 1
                guardados += 1

            else:
                # DESCONOCIDO → creamos/llenamos "Desconocidos" (queda al final)
                if "Desconocidos" not in wb.sheetnames:
                    hoja_desc = wb.create_sheet("Desconocidos")
                    _poner_encabezados(hoja_desc)
                else:
                    hoja_desc = wb["Desconocidos"]
                hoja_desc.append([
                    ahora.strftime("%d/%m/%Y"),
                    ahora.strftime("%H:%M:%S"),
                    nombre,
                    tiempo_mmss or "",
                    int(tiempo_seg) if isinstance(tiempo_seg, (int, float)) else "",
                    vel_kmh if vel_kmh is not None else "",
                    dist_km if dist_km is not None else ""
                ])
                guardados += 1

        except IllegalCharacterError:
            continue
        except Exception:
            continue

    # Reordenar hojas conforme a la política y guardar
    try:
        _ordenar_hojas_watchlist()
        safe_save_workbook(wb, archivo_excel)
    except Exception:
        pass

    return guardados, u, i

__all__ = [
    "wb",
    "archivo_excel",
    "set_workbook",
    "ruta_log_txt",
    "atomic_save_workbook",
    "safe_save_workbook",
    "guardar_tramos",
]
